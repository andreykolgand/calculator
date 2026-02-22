from __future__ import annotations

import json
from dataclasses import dataclass
from typing import List, Tuple, Dict
from io import BytesIO
from datetime import datetime

from flask import Flask, render_template, request, send_file


app = Flask(__name__)


@dataclass
class PaymentRow:
    month: int
    payment: float
    principal: float
    interest: float
    balance: float
    prepayment: float = 0.0


@dataclass
class Prepayment:
    month: int
    amount: float


def parse_number(value: str, default: float = 0.0) -> float:
    """Parse a number from a string that may contain spaces or separators."""
    if value is None:
        return default
    cleaned = "".join(ch for ch in value if (ch.isdigit() or ch in ",."))
    if not cleaned:
        return default
    # Replace comma with dot to handle "12,5"
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return default


def format_rubles(value: float, with_decimals: bool = False) -> str:
    """
    Format numbers with space as thousands separator (e.g. 1 234 567.89).
    """
    if with_decimals:
        s = f"{value:,.2f}"
    else:
        s = f"{int(round(value)):,.0f}"
    # Convert 1,234,567.89 -> 1 234 567.89
    s = s.replace(",", " ")
    return s


def calculate_schedule(
    total_amount: float,
    down_payment: float,
    annual_rate: float,
    years: int,
    prepayments: List[Prepayment] = None,
    reduce_payment: bool = True,
) -> Tuple[float, float, float, float, float, List[PaymentRow]]:
    """
    Calculate mortgage schedule using annuity payments with optional prepayments.

    Args:
        total_amount: Total loan amount
        down_payment: Down payment amount
        annual_rate: Annual interest rate (percentage)
        years: Loan term in years
        prepayments: List of prepayments (month, amount)
        reduce_payment: If True, reduce payment amount; if False, reduce term

    Returns:
        principal, monthly_payment, total_paid, total_interest, overpayment, schedule
    """
    if prepayments is None:
        prepayments = []

    principal = max(total_amount - down_payment, 0)
    months = max(years * 12, 0)

    if principal <= 0 or months <= 0:
        return 0, 0, 0, 0, 0, []

    monthly_rate = annual_rate / 12.0 / 100.0

    # Создаем словарь досрочных платежей по месяцам
    prepayment_dict: Dict[int, float] = {}
    for prep in prepayments:
        month = int(prep.month)
        if month > 0:
            if month in prepayment_dict:
                prepayment_dict[month] += prep.amount
            else:
                prepayment_dict[month] = prep.amount

    # Изначальный аннуитетный платеж
    if monthly_rate == 0:
        initial_monthly_payment = principal / months
    else:
        factor = (1 + monthly_rate) ** months
        initial_monthly_payment = principal * monthly_rate * factor / (factor - 1)

    balance = principal
    schedule: List[PaymentRow] = []
    total_interest = 0.0
    current_monthly_payment = initial_monthly_payment
    month = 1
    max_months = months * 3  # Защита от бесконечного цикла

    while balance > 0.01 and month <= max_months:
        if monthly_rate == 0:
            interest_payment = 0.0
        else:
            interest_payment = balance * monthly_rate

        principal_payment = current_monthly_payment - interest_payment
        prepayment_amount = prepayment_dict.get(month, 0.0)

        # Применяем основной платеж
        balance -= principal_payment

        # Применяем досрочный платеж (после основного платежа)
        if prepayment_amount > 0:
            balance -= prepayment_amount
            prepayment_amount = max(prepayment_amount, 0)

        if balance < 0:
            # Если баланс стал отрицательным, корректируем
            if prepayment_amount > 0:
                prepayment_amount += balance
                prepayment_amount = max(prepayment_amount, 0)
            balance = 0

        total_interest += interest_payment

        schedule.append(
            PaymentRow(
                month=month,
                payment=current_monthly_payment,
                principal=max(principal_payment, 0),
                interest=max(interest_payment, 0),
                balance=max(balance, 0),
                prepayment=prepayment_amount,
            )
        )

        # Если есть досрочный платеж и остался долг, пересчитываем график
        if prepayment_amount > 0 and balance > 0.01:
            remaining_months = max(months - month, 1)

            if reduce_payment:
                # Уменьшаем платеж: пересчитываем аннуитет с новым остатком и оставшимся сроком
                if monthly_rate == 0:
                    current_monthly_payment = balance / remaining_months
                else:
                    factor = (1 + monthly_rate) ** remaining_months
                    if factor > 1:
                        current_monthly_payment = balance * monthly_rate * factor / (factor - 1)
                    else:
                        current_monthly_payment = balance / remaining_months
            # Если reduce_payment=False, платеж остается прежним, срок уменьшится автоматически

        month += 1

        # Если срок истек и баланс не погашен, продолжаем до полного погашения
        if month > months and balance > 0.01 and not reduce_payment:
            continue

        if balance <= 0.01:
            break

    total_paid = sum(row.payment + row.prepayment for row in schedule)
    # Переплата = только проценты (не включает досрочные платежи)
    overpayment = total_interest

    return principal, initial_monthly_payment, total_paid, total_interest, overpayment, schedule


@app.route("/", methods=["GET", "POST"])
def index():
    # Значения по умолчанию
    loan_amount_raw = "5 000 000"
    down_payment_raw = "1 000 000"
    years_raw = "20"
    rate_raw = "12"

    principal = 0.0
    monthly_payment = 0.0
    total_paid = 0.0
    total_interest = 0.0
    overpayment = 0.0
    schedule: List[PaymentRow] = []
    schedule_for_export = []
    down_percent = 0.0
    min_income = 0.0
    prepayments_data = "[]"
    prepayment_strategy = "reduce_payment"

    if request.method == "POST":
        loan_amount_raw = request.form.get("loan_amount", loan_amount_raw)
        down_payment_raw = request.form.get("down_payment", down_payment_raw)
        years_raw = request.form.get("years", years_raw)
        rate_raw = request.form.get("rate", rate_raw)
        prepayments_data = request.form.get("prepayments_data", "[]")
        prepayment_strategy = request.form.get("prepayment_strategy", "reduce_payment")

        loan_amount = parse_number(loan_amount_raw)
        down_payment = parse_number(down_payment_raw)
        years = int(parse_number(years_raw))
        rate = parse_number(rate_raw)

        # Ограничения по сумме кредита и ставке
        loan_amount = max(min(loan_amount, 40_000_000), 500_000)
        rate = max(min(rate, 40.0), 0.0)

        # Ограничения по размеру первоначального взноса (15–70%)
        if loan_amount > 0:
            min_dp = loan_amount * 0.15
            max_dp = loan_amount * 0.7
            down_payment = max(min(down_payment, max_dp), min_dp)
        else:
            down_payment = 0.0

        # Парсим досрочные платежи
        prepayments: List[Prepayment] = []
        try:
            prepayments_json = json.loads(prepayments_data)
            for item in prepayments_json:
                month = int(item.get("month", 0))
                amount = parse_number(str(item.get("amount", 0)))
                if month > 0 and amount > 0:
                    prepayments.append(Prepayment(month=month, amount=amount))
        except (json.JSONDecodeError, ValueError, TypeError):
            prepayments = []

        reduce_payment = prepayment_strategy == "reduce_payment"

        (
            principal,
            monthly_payment,
            total_paid,
            total_interest,
            overpayment,
            schedule,
        ) = calculate_schedule(
            loan_amount, down_payment, rate, years, prepayments, reduce_payment
        )

        if loan_amount > 0:
            down_percent = max(
                min(down_payment / loan_amount * 100.0, 100.0),
                0.0,
            )
        else:
            down_percent = 0.0

        # Расчет минимального необходимого дохода
        # Банки обычно требуют, чтобы платеж не превышал 40% от дохода
        # Минимальный доход = ежемесячный платеж / 0.4
        if monthly_payment > 0:
            min_income = monthly_payment / 0.4
        else:
            min_income = 0.0

        # Нормализуем отображение ввода
        loan_amount_raw = format_rubles(loan_amount)
        down_payment_raw = format_rubles(down_payment)
        years_raw = str(years)
        rate_raw = str(rate).replace(".", ",")

        # Подготавливаем данные графика для экспорта
        schedule_for_export = [
            {
                "month": row.month,
                "payment": row.payment,
                "principal": row.principal,
                "interest": row.interest,
                "prepayment": row.prepayment,
                "balance": row.balance,
            }
            for row in schedule
        ]

    return render_template(
        "index.html",
        loan_amount_raw=loan_amount_raw,
        down_payment_raw=down_payment_raw,
        years_raw=years_raw,
        rate_raw=rate_raw,
        principal=principal,
        monthly_payment=monthly_payment,
        total_paid=total_paid,
        total_interest=total_interest,
        overpayment=overpayment,
        schedule=schedule,
        schedule_for_export=schedule_for_export,
        down_percent=down_percent,
        min_income=min_income,
        prepayments_data=prepayments_data,
        prepayment_strategy=prepayment_strategy,
        format_rubles=format_rubles,
    )


def _get_export_params():
    """Читает параметры экспорта из JSON body или из form (для совместимости)."""
    if request.content_type and "application/json" in request.content_type:
        data = request.get_json(silent=True) or {}
        schedule_data = data.get("schedule_data", [])
        return {
            "schedule_data": schedule_data if isinstance(schedule_data, list) else [],
            "loan_amount": str(data.get("loan_amount", "")),
            "down_payment": str(data.get("down_payment", "")),
            "years": str(data.get("years", "")),
            "rate": str(data.get("rate", "")),
            "principal": str(data.get("principal", "0")),
            "monthly_payment": str(data.get("monthly_payment", "0")),
            "total_paid": str(data.get("total_paid", "0")),
            "overpayment": str(data.get("overpayment", "0")),
        }
    schedule_json = request.form.get("schedule_data", "[]")
    try:
        schedule_data = json.loads(schedule_json)
    except (json.JSONDecodeError, TypeError):
        schedule_data = []
    return {
        "schedule_data": schedule_data if isinstance(schedule_data, list) else [],
        "loan_amount": request.form.get("loan_amount", ""),
        "down_payment": request.form.get("down_payment", ""),
        "years": request.form.get("years", ""),
        "rate": request.form.get("rate", ""),
        "principal": request.form.get("principal", "0"),
        "monthly_payment": request.form.get("monthly_payment", "0"),
        "total_paid": request.form.get("total_paid", "0"),
        "overpayment": request.form.get("overpayment", "0"),
    }


@app.route("/export_excel", methods=["POST"])
def export_excel():
    """Экспорт графика платежей в Excel (.xlsx) — только таблица"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return export_csv()

    params = _get_export_params()
    schedule_data = params["schedule_data"]
    loan_amount = params["loan_amount"]
    down_payment = params["down_payment"]
    years = params["years"]
    rate = params["rate"]
    principal = params["principal"]
    monthly_payment = params["monthly_payment"]
    total_paid = params["total_paid"]
    overpayment = params["overpayment"]

    wb = Workbook()
    ws = wb.active
    ws.title = "График платежей"

    header_fill = PatternFill(start_color="D9B15F", end_color="D9B15F", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = "График платежей по ипотечному кредиту"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    ws[f"A{row}"] = "Параметры кредита:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws[f"A{row}"] = f"Сумма кредита: {loan_amount} ₽"
    row += 1
    ws[f"A{row}"] = f"Первоначальный взнос: {down_payment} ₽"
    row += 1
    ws[f"A{row}"] = f"Срок кредита: {years} лет"
    row += 1
    ws[f"A{row}"] = f"Процентная ставка: {rate}% годовых"
    row += 2

    ws[f"A{row}"] = "Результаты расчета:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws[f"A{row}"] = f"Основной долг: {format_rubles(float(principal), False)} ₽"
    row += 1
    ws[f"A{row}"] = f"Ежемесячный платёж: {format_rubles(float(monthly_payment), True)} ₽"
    row += 1
    ws[f"A{row}"] = f"Сумма всех платежей: {format_rubles(float(total_paid), True)} ₽"
    row += 1
    ws[f"A{row}"] = f"Проценты банку (переплата): {format_rubles(float(overpayment), True)} ₽"
    row += 2

    headers = ["Месяц", "Платёж, ₽", "Из них тело, ₽", "Из них проценты, ₽", "Досрочный платёж, ₽", "Остаток долга, ₽"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    row += 1

    for item in schedule_data:
        month_val = item.get("month", 0)
        try:
            payment_val = float(item.get("payment", 0) or 0)
            principal_val = float(item.get("principal", 0) or 0)
            interest_val = float(item.get("interest", 0) or 0)
            prepayment_val = float(item.get("prepayment", 0) or 0)
            balance_val = float(item.get("balance", 0) or 0)
        except (TypeError, ValueError):
            payment_val = principal_val = interest_val = prepayment_val = balance_val = 0.0
        ws.cell(row=row, column=1, value=month_val).alignment = center_align
        ws.cell(row=row, column=2, value=format_rubles(payment_val, True)).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=format_rubles(principal_val, True)).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4, value=format_rubles(interest_val, True)).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5, value=format_rubles(prepayment_val, True) if prepayment_val > 0 else "-").alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=format_rubles(balance_val, True)).alignment = Alignment(horizontal="right")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        row += 1

    column_widths = [12, 18, 18, 20, 22, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"График_платежей_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def export_csv():
    """Резервный экспорт в CSV если openpyxl не установлен"""
    params = _get_export_params()
    schedule_data = params["schedule_data"]

    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output, delimiter=";")

    # Заголовки
    writer.writerow(["Месяц", "Платёж, ₽", "Из них тело, ₽", "Из них проценты, ₽", "Досрочный платёж, ₽", "Остаток долга, ₽"])

    # Данные
    for item in schedule_data:
        prepayment = item.get("prepayment", 0)
        writer.writerow([
            item.get("month", 0),
            format_rubles(item.get("payment", 0), True),
            format_rubles(item.get("principal", 0), True),
            format_rubles(item.get("interest", 0), True),
            format_rubles(prepayment, True) if prepayment > 0 else "-",
            format_rubles(item.get("balance", 0), True),
        ])

    output.seek(0)
    filename = f"График_платежей_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return send_file(
        BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv; charset=utf-8-sig",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=True)
